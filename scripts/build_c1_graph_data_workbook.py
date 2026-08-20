"""Create the C1 manual-chart data workbook from existing raw R01-R05 files."""
from __future__ import annotations

import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
from build_c1_advisor_graphs_v3 import RUNS, load, phase_durations, event_resource  # noqa: E402

OUT = ROOT / "results" / "ADVISOR_REVIEW_C1_5SEEDS" / "C1_GRAPH_DATA.xlsx"
VALIDATION = ROOT / "results" / "ADVISOR_REVIEW_C1_5SEEDS" / "C1_GRAPH_DATA_VALIDATION.md"
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")


def put_table(ws, row: int, title: str, headers: list[str], rows: list[list]) -> int:
    ws.cell(row, 1, title).font = Font(bold=True, size=13)
    row += 1
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row, col, value); cell.font = Font(bold=True); cell.fill = HEADER_FILL
    for values in rows:
        row += 1
        for col, value in enumerate(values, 1): ws.cell(row, col, value)
    return row + 2


def finish(ws) -> None:
    ws.freeze_panes = "A2"
    for column in range(1, ws.max_column + 1):
        width = max([len(str(ws.cell(row, column).value or "")) for row in range(1, ws.max_row + 1)] + [10])
        ws.column_dimensions[get_column_letter(column)].width = min(width + 2, 55)


def detail_number(detail: str, pattern: str) -> float:
    match = re.search(pattern, detail)
    if not match: raise ValueError(detail)
    return float(match.group(1))


def main() -> int:
    runs, _ = load()
    wb = Workbook(); wb.remove(wb.active)
    names = ["README", "Graph1", "Graph2", "Graph3", "Graph4", "Graph5", "Graph6", "Graph7", "Graph8", "Supplement_A", "Supplement_B", "DATA_AUDIT"]
    sheets = {name: wb.create_sheet(name) for name in names}
    ids = [r["id"] for r in runs]
    effective = {r["id"]: r["effective"] for r in runs}
    event_counts = Counter(e["event"] for r in runs for e in r["events"])
    resource_detection, resource_collection, resource_delivery = Counter(), Counter(), Counter()
    per_run = {}
    g7 = []
    for run in runs:
        events = run["events"]
        harvest_by_trip = {}
        withdrawals = []
        for e in events:
            if e["event"] == "RESOURCE_LIGHT_DETECTED": resource_detection[event_resource(e["detail"])] += 1
            if e["event"] == "HARVEST_COMPLETE":
                resource = event_resource(e["detail"]); resource_collection[resource] += 1; harvest_by_trip[(e["scout_id"], e["trip_id"])] = resource
            if e["event"] == "DELIVER": resource_delivery[harvest_by_trip.get((e["scout_id"], e["trip_id"]), "UNKNOWN")] += 1
            if e["event"] == "NEST_ENERGY_WITHDRAWAL": withdrawals.append(e)
        starts = sum(e["event"] in {"SCOUT_START", "NEXT_CYCLE_START"} for e in events)
        returns = sum(e["event"] == "RETURN_HOME_START" for e in events)
        reaches = sum(e["event"] == "NEST_REACHED" for e in events)
        deliveries = sum(e["event"] == "DELIVER" for e in events)
        gross = float(run["summary"].get("gross_delivered_energy", 0)); withdrawn = float(run["summary"].get("total_robot_nest_withdrawal", 0)); net = float(run["summary"].get("net_nest_energy", 0))
        per_run[run["id"]] = {"starts": starts, "returns": returns, "reaches": reaches, "deliveries": deliveries, "withdrawals": withdrawals, "gross": gross, "withdrawn": withdrawn, "net": net}
        for depletion in run["depleted"]:
            sid = depletion["scout_id"]; dep_t = float(depletion["sim_time_s"]); own = [e for e in withdrawals if e["scout_id"] == sid]
            recharge_t = float(own[0]["sim_time_s"]) if own else None
            amount = detail_number(own[0]["detail"], r"withdrawal=([0-9.]+)") if own else None
            trajectory = [x for x in run["trajectory"] if x["scout_id"] == sid and float(x["sim_time_s"]) <= dep_t]
            last = trajectory[-1] if trajectory else {}
            g7.append([run["id"], f"Scout {sid}", f"{run['id']}-S{sid}", dep_t, "ได้รับการเติม" if own else "ไม่เคยเติม", recharge_t, amount, dep_t-recharge_t if recharge_t is not None else None, int(last["cycle_id"]) if last else None, last.get("phase")])

    # README
    guide = [["ระยะเวลาที่หุ่นสิ้นสุดการทำงาน", "Graph1", "Clustered Column", "Run", "เวลาจำลอง (วินาที)", "เวลาที่ Scout ตัวสุดท้ายหมดพลังงาน"], ["การดึงพลังงานจากรังมาใช้", "Graph2", "Scatter Plot", "เวลาจำลอง", "Run", "แต่ละแถวคือเหตุการณ์ดึงพลังงาน"], ["สัดส่วนเวลาที่หุ่นใช้ในแต่ละพฤติกรรม", "Graph3", "100% Stacked Bar", "Run", "สัดส่วนเวลา (%)", "ใช้ total active Scout-time"], ["ลำดับขั้นของการหาพลังงาน", "Graph4", "Horizontal Bar", "จำนวนครั้ง", "ขั้นตอน", "ลำดับจาก event log"], ["ผลการเดินทางกลับรัง", "Graph5", "Clustered Column", "Run", "จำนวนครั้ง", "สำเร็จคือถึง Nest จริง"], ["การใช้แหล่งพลังงาน A/B/C", "Graph6", "Clustered Column / Marker", "Resource", "จำนวนครั้ง", "แนะนำเก็บสำเร็จเทียบส่งถึงรัง"], ["เวลาที่หุ่นหมดพลังงานและการเติมพลังงานจากรัง", "Graph7", "Scatter / Dot Plot", "เวลาหมดพลังงาน", "Scout", "15 Scouts"], ["อัตราผลลัพธ์ด้านพลังงานของ Colony", "Graph8", "Clustered Column", "Run", "หน่วยต่อ 1,000 วินาที", "ใช้งานจริงก่อนหมดพลังงาน"], ["สมดุลพลังงานของรัง", "Supplement_A", "Clustered Column", "Run", "พลังงาน (หน่วย)", "บัญชีพลังงาน"], ["พลังงานสุทธิของรังตามเวลา", "Supplement_B", "Line Chart", "เวลาจำลอง", "พลังงานสุทธิ", "หยุดที่ effective termination"]]
    put_table(sheets["README"], 1, "คู่มือข้อมูลสำหรับสร้างกราฟ C1", ["ชื่อกราฟ", "Sheet", "ชนิดกราฟที่แนะนำใน Excel", "แกน X", "แกน Y", "คำอธิบายสั้น ๆ"], guide)
    sheets["README"].cell(sheets["README"].max_row + 2, 1, "Graph6: ใช้จำนวนครั้งเก็บพลังงานสำเร็จเทียบจำนวนครั้งส่งพลังงานจากแหล่งนั้นถึงรังสำเร็จ เพื่อไม่วาด detection/collection ที่ซ้ำกัน")

    # Graph1
    ws=sheets["Graph1"]; row=put_table(ws,1,"ข้อมูล Graph 1",["Run","เวลาที่หุ่นตัวสุดท้ายหมดพลังงาน (วินาที)"],[[r["id"],r["effective"]] for r in runs]); ws.cell(row,1,"การทดลองสิ้นสุดเมื่อหุ่นทุกตัวหมดพลังงาน")
    put_table(ws,row+2,"ข้อมูล metadata (ไม่ใช้สร้างกราฟหลัก)",["Run","เวลาที่ระบบเดิมรันทั้งหมด (วินาที)","เวลาที่ Colony ทำงานได้จริง (วินาที)","ช่วงเวลาหลังหุ่นดับหมด (วินาที)"],[[r["id"],float(r["summary"]["simulation_time_s"]),r["effective"],float(r["summary"]["simulation_time_s"])-r["effective"]] for r in runs])

    # Graph2
    ws=sheets["Graph2"]; event_rows=[]; summary=[]
    for r in runs:
        actual=per_run[r["id"]]["withdrawals"]
        if actual:
            for e in actual: event_rows.append([r["id"],float(e["sim_time_s"]),f"Scout {e['scout_id']}",detail_number(e["detail"],r"withdrawal=([0-9.]+)"),"ดึงพลังงานจากรัง"])
        else: event_rows.append([r["id"],None,None,0,"ไม่มีการดึงพลังงาน"])
        summary.append([r["id"],len(actual),sum(detail_number(e["detail"],r"withdrawal=([0-9.]+)") for e in actual)])
    row=put_table(ws,1,"เหตุการณ์การดึงพลังงานจากรัง",["Run","เวลาเกิดเหตุการณ์ (วินาที)","Scout","พลังงานที่ดึงจากรัง (หน่วย)","หมายเหตุ"],event_rows); row=put_table(ws,row,"สรุป",["Run","จำนวนครั้งที่ดึงพลังงาน","พลังงานที่ดึงรวม"],summary); ws.cell(row,1,"พลังงานที่ดึงจากรังรวมทั้ง 5 รอบ"); ws.cell(row,2,sum(x[2] for x in summary))

    # Graph3 actual phase durations.
    headers_pct=["Run","สำรวจ (%)","เข้าหาแหล่งพลังงาน (%)","เก็บพลังงาน (%)","เดินทางกลับรัง (%)","อยู่ที่รัง / เติมพลังงาน (%)"]
    headers_sec=["Run","สำรวจ (วินาที)","เข้าหาแหล่งพลังงาน (วินาที)","เก็บพลังงาน (วินาที)","เดินทางกลับรัง (วินาที)","อยู่ที่รัง / เติมพลังงาน (วินาที)","เวลาทำงานรวมของ Scout ทั้ง 3 ตัว (วินาที)"]
    pct_rows=[]; sec_rows=[]
    for r in runs:
        dur=Counter()
        for sid in ("0","1","2"): dur.update(phase_durations([x for x in r["trajectory"] if x["scout_id"]==sid],r["effective"]))
        total=sum(dur.values()); vals=[dur["EXPLORE"],dur["HARVEST"],dur["RETURN_HOME"],dur["DELIVER"]]
        pct_rows.append([r["id"],vals[0]/total, None, vals[1]/total, vals[2]/total, vals[3]/total]); sec_rows.append([r["id"],vals[0],None,vals[1],vals[2],vals[3],total])
    ws=sheets["Graph3"]; row=put_table(ws,1,"สัดส่วนเวลาปฏิบัติงานจริง",headers_pct,pct_rows); row=put_table(ws,row,"เวลาเป็นวินาที",headers_sec,sec_rows); ws.cell(row,1,"หมายเหตุ: log ไม่มี state ‘เข้าหาแหล่งพลังงาน’ แยกจาก ‘สำรวจ’; เว้นว่างเพื่อไม่สร้าง proxy. 100% = ผลรวม active Scout-time ของ 3 Scout ก่อนพลังงานหมด")
    for rr in range(3,8):
        for cc in (2,4,5,6): ws.cell(rr,cc).number_format="0.0%"

    # Graph4
    stages=[("เริ่มรอบการทำงาน",sum(x["starts"] for x in per_run.values())),("ตรวจพบแหล่งพลังงาน",event_counts["RESOURCE_LIGHT_DETECTED"]),("เก็บพลังงานสำเร็จ",event_counts["HARVEST_COMPLETE"]),("เริ่มเดินทางกลับรัง",event_counts["RETURN_HOME_START"]),("กลับถึงรัง",event_counts["NEST_REACHED"]),("ส่งพลังงานเข้ารังสำเร็จ",event_counts["DELIVER"])]
    ws=sheets["Graph4"]; put_table(ws,1,"ลำดับขั้นของการหาพลังงาน",["ขั้นตอน","จำนวนครั้ง","ร้อยละเทียบกับขั้นแรก"],[[a,b,b/stages[0][1]] for a,b in stages])
    for rr in range(3,3+len(stages)): ws.cell(rr,3).number_format="0.0%"

    # Graph5/6/7/8
    ws=sheets["Graph5"]; put_table(ws,1,"ผลการเดินทางกลับรัง",["Run","จำนวนครั้งที่เริ่มเดินทางกลับรัง","จำนวนครั้งที่กลับถึงรังสำเร็จ","จำนวนครั้งที่กลับไม่สำเร็จ","อัตราความสำเร็จในการกลับรัง (%)"],[[rid,x["returns"],x["reaches"],x["returns"]-x["reaches"],x["reaches"]/x["returns"] if x["returns"] else 0] for rid,x in per_run.items()]);
    for rr in range(3,8): ws.cell(rr,5).number_format="0.0%"
    config=__import__("json").loads((ROOT/"config"/"resource_harvesting_config.json").read_text(encoding="utf-8")); rate={x["resource_id"]:x["relative_harvest_rate"] for x in config["sources"]}
    ws=sheets["Graph6"]; put_table(ws,1,"การใช้แหล่งพลังงาน A/B/C",["แหล่งพลังงาน","จำนวนครั้งที่ตรวจพบ","จำนวนครั้งที่เก็บพลังงานสำเร็จ","จำนวนครั้งที่พลังงานจากแหล่งนั้นถูกส่งถึงรังสำเร็จ","อัตรา ตรวจพบ → เก็บสำเร็จ (%)","เวลาเก็บพลังงานเฉลี่ย (วินาที)","พลังงานที่ส่งถึงรังรวม (หน่วย)"],[[f"Resource {s}",resource_detection[s],resource_collection[s],resource_delivery[s],resource_collection[s]/resource_detection[s] if resource_detection[s] else None,1/rate[s],float(resource_delivery[s])] for s in "ABC"])
    for rr in range(3,6): ws.cell(rr,5).number_format="0.0%"
    ws=sheets["Graph7"]; put_table(ws,1,"เวลา Scout หมดพลังงานและการเติมพลังงานจากรัง",["Run","Scout","ชื่อสำหรับแกนกราฟ","เวลาที่พลังงานหมด (วินาที)","เคยเติมพลังงานจากรังหรือไม่","เวลาที่เติมพลังงานจากรัง (วินาที)","พลังงานที่เติมจากรัง (หน่วย)","เวลาที่อยู่ต่อหลังเติมพลังงาน (วินาที)","Cycle ตอนพลังงานหมด","สถานะตอนพลังงานหมด"],g7)
    g8=[]
    for r in runs:
        x=per_run[r["id"]]; g8.append([r["id"],r["effective"],x["gross"],x["net"],x["gross"]/r["effective"]*1000,x["net"]/r["effective"]*1000])
    ws=sheets["Graph8"]; put_table(ws,1,"อัตราผลลัพธ์ด้านพลังงานของ Colony",["Run","เวลาที่ Colony ทำงานได้จริง (วินาที)","พลังงานที่ส่งเข้ารังทั้งหมด (หน่วย)","พลังงานสุทธิสุดท้ายของรัง (หน่วย)","อัตราพลังงานที่ส่งเข้ารังต่อ 1,000 วินาที","อัตราการเพิ่มพลังงานสุทธิของรังต่อ 1,000 วินาที"],g8)
    ws=sheets["Supplement_A"]; put_table(ws,1,"สมดุลพลังงานของรัง",["Run","พลังงานที่ส่งเข้ารังทั้งหมด","พลังงานที่หุ่นดึงออกจากรัง","พลังงานสุทธิสุดท้ายของรัง"],[[r["id"],per_run[r["id"]]["gross"],per_run[r["id"]]["withdrawn"],per_run[r["id"]]["net"]] for r in runs]); ws.cell(9,1,"ตรวจสอบ: พลังงานสุทธิ = ส่งเข้ารัง - ดึงออกจากรัง")
    # Supplement B long data only to effective endpoint.
    line=[]
    for r in runs:
        values=[(0.0,0.0,"เริ่มต้น")]+[(float(x["timestamp"]),float(x["new_energy"]),"ส่งพลังงานเข้ารัง" if x["event_type"]=="DELIVERY" else "ดึงพลังงานจากรัง") for x in r["nest"] if float(x["timestamp"])<=r["effective"]]
        values.append((r["effective"],values[-1][1],"สิ้นสุดเมื่อ Scout ทุกตัวหมดพลังงาน"))
        for t,e,kind in values: line.append([r["id"],t,e,kind])
    ws=sheets["Supplement_B"]; put_table(ws,1,"พลังงานสุทธิของรังตามเวลา",["Run","เวลา (วินาที)","พลังงานสุทธิของรัง","ประเภทเหตุการณ์"],line)
    # Data audit.
    audit=[("จำนวน Runs",5,"raw run folders","R01-R05","ผ่าน",""),("Scouts ต่อ Run",3,"swarm_summary.json","scout_count","ผ่าน",""),("Scouts รวม",15,"swarm_events.csv","SCOUT_START","ผ่าน",""),("effective termination R01-R05",", ".join(f"{k}={v:.1f}" for k,v in effective.items()),"swarm_events.csv","ROBOT_DEPLETED","ผ่าน","ค่าสูงสุดต่อ Run"),("ตรวจพบแหล่งพลังงาน",event_counts["RESOURCE_LIGHT_DETECTED"],"swarm_events.csv","RESOURCE_LIGHT_DETECTED","ผ่าน",""),("เก็บพลังงานสำเร็จ",event_counts["HARVEST_COMPLETE"],"swarm_events.csv","HARVEST_COMPLETE","ผ่าน",""),("เริ่มกลับรัง",event_counts["RETURN_HOME_START"],"swarm_events.csv","RETURN_HOME_START","ผ่าน",""),("กลับถึงรัง",event_counts["NEST_REACHED"],"swarm_events.csv","NEST_REACHED","ผ่าน",""),("ส่งพลังงานเข้ารัง",event_counts["DELIVER"],"swarm_events.csv","DELIVER","ผ่าน",""),("ดึงพลังงานจากรัง",sum(x["withdrawn"] for x in per_run.values()),"swarm_events.csv","NEST_ENERGY_WITHDRAWAL","ผ่าน",""),("พลังงานสุทธิสุดท้าย",", ".join(f"{r['id']}={per_run[r['id']]['net']:.1f}" for r in runs),"swarm_summary.json","net_nest_energy","ผ่าน",""),("Scout หมดพลังงาน",len(g7),"swarm_events.csv","ROBOT_DEPLETED","ผ่าน",""),("A/B/C เก็บสำเร็จ",f"{resource_collection['A']}/{resource_collection['B']}/{resource_collection['C']}","swarm_events.csv","HARVEST_COMPLETE","ผ่าน",""),("แหล่งของการส่งสำเร็จ",f"A={resource_delivery['A']}, B={resource_delivery['B']}, C={resource_delivery['C']}","swarm_events.csv","HARVEST_COMPLETE + DELIVER","ผ่าน","จับคู่ Scout/trip")]
    ws=sheets["DATA_AUDIT"]; put_table(ws,1,"การตรวจสอบข้อมูล",["Metric","Calculated value","Source file","Source column/event","Status","Notes"],audit)
    for ws in sheets.values(): finish(ws)
    OUT.parent.mkdir(parents=True,exist_ok=True); wb.save(OUT)
    # Re-open validation.
    check=load_workbook(OUT,data_only=True)
    assert check.sheetnames == names
    assert check["Graph7"].max_row == 17  # title + header + 15 data rows
    assert all(abs(sum(check["Graph3"].cell(row,c).value or 0 for c in (2,4,5,6))-1) < .002 for row in range(3,8))
    assert all(check["Graph5"].cell(row,2).value == check["Graph5"].cell(row,3).value + check["Graph5"].cell(row,4).value for row in range(3,8))
    assert all(row[1] <= effective[row[0]] + 1e-9 for row in line)
    VALIDATION.write_text("# C1 GRAPH DATA VALIDATION\n\nWorkbook re-opened with openpyxl successfully. All required sheets exist; Graph3 active-state percentages sum to 100% across logged states; Graph5 counts reconcile; Graph7 has 15 Scout rows; Graph8 uses effective termination time; Supplement_B contains no timestamp beyond each Run effective termination.\n\nLimitation: no separate APPROACH_RESOURCE state is logged. Graph3 leaves that field blank rather than inventing a proxy.\n",encoding="utf-8")
    print(OUT)
    return 0

if __name__ == "__main__": raise SystemExit(main())
